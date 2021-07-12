import time
import tkinter as tk
import pandas as pd
import tkinter.messagebox 
from openpyxl import *
import datetime
from tkinter import *

wb = load_workbook('Database.xlsx') 

sheet = wb.active 

data = []
root=Tk()
root.title("Generate Employee Salary Report")
root.geometry('1000x550+0+0')
root.configure(background="light yellow")

Tops=Frame(root,width=1200,height=25,bd=4,bg="green")
Tops.pack(side=TOP)
 
f1=Frame(root,width=300,height=150,bd=4,bg="red")
f1.pack(side=LEFT)
f2=Frame(root,width=150,height=350,bd=4,bg="light yellow")
f2.pack(side=RIGHT)
 
fla=Frame(f1,width=300,height=100,bd=4,bg="light yellow")
fla.pack(side=TOP)
flb=Frame(f1,width=150,height=300,bd=4,bg="red")
flb.pack(side=TOP)
 
lblinfo=Label(Tops,font=('arial',30,'bold'),text="Generate Employee Salary Report ",bd=10,bg="light yellow",fg="green")
lblinfo.grid(row=0,column=0)
sheet.column_dimensions['A'].width = 30
sheet.column_dimensions['B'].width = 20
sheet.column_dimensions['C'].width = 20
sheet.column_dimensions['D'].width = 20
sheet.column_dimensions['E'].width = 20
sheet.column_dimensions['F'].width = 20
sheet.column_dimensions['G'].width = 20
sheet.column_dimensions['I'].width = 20
sheet.column_dimensions['J'].width = 20
sheet.cell(row=1, column=1).value = "Name"
sheet.cell(row=1, column=2).value = "Address"
sheet.cell(row=1, column=3).value = "Employer"
sheet.cell(row=1, column=4).value = "EmployeeId"
sheet.cell(row=1, column=5).value = "HoursWorked"
sheet.cell(row=1, column=6).value = "NetPayable"
sheet.cell(row=1, column=7).value = "wageshour"
sheet.cell(row=1, column=8).value = "Taxable"
sheet.cell(row=1, column=9).value = "Payable"
def exit():
  exit=tkinter.messagebox.askyesno("Employee system","Do you want to exit the system")
  if exit>0:
    root.destroy()
    return
def saveinfo():
    valor1 = Name.get()
    valor2 = Address.get()
    valor3 = Employer.get()
    valor4 = EmployeeId.get()
    valor5 = HoursWorked.get()
    valor6 = NetPayable.get()
    valor7 = wageshour.get()
    valor8 = Taxable.get()
    valor9 = Payable.get()



    data.append([valor1, valor2, valor3, valor4, valor5, valor6, valor7, valor8,valor9 ])
    print(data,"\n")



def export():
  current_row = sheet.max_row 
  current_column = sheet.max_column
  sheet.cell(row=current_row + 1, column=1).value = Name.get() 
  sheet.cell(row=current_row + 1, column=2).value = Address.get() 
  sheet.cell(row=current_row + 1, column=3).value = Employer.get() 
  sheet.cell(row=current_row + 1, column=4).value = EmployeeId.get()
  sheet.cell(row=current_row + 1, column=5).value = HoursWorked.get() 
  sheet.cell(row=current_row + 1, column=6).value = NetPayable.get() 
  sheet.cell(row=current_row + 1, column=7).value = wageshour.get() 
  sheet.cell(row=current_row + 1, column=8).value = EmployeeId.get() 
  sheet.cell(row=current_row + 1, column=9).value = Payable.get() 
    
    

    # save the file 
  wb.save('Database.xlsx') 
  enterinfo()
  saveinfo()
  reset()


def reset():
  Name.set("")
  Address.set("")
  HoursWorked.set("")
  wageshour.set("")
  Payable.set("")
  Taxable.set("")
  NetPayable.set("")
  GrossPayable.set("")
  OverTimeBonus.set("")
  Employer.set("")
  EmployeeId.set("")

#txtpayslip.delete("1.0",END)
''' 
def findRecord(eno,en,da,sl,edf):
    e = eno.get()
    csvfile=csv.reader(open("excel.xlsx","r"))

    for row in csvfile:
      if e==row[3]:
        print(row)
def find():
  eno=StringVar()
  en=StringVar()
  da=StringVar()
  sl=StringVar()


  searchForm = Toplevel(root)
  searchForm.title("...:::Search Employee")
  searchForm.geometry("450x350+400+220")
  searchForm.resizable(False,False)
  lb1=Label(searchForm,text="SEARCH EMPLOYEE FORM",bg='black',fg='white',font=('verdana',12,'bold'),width=30)
  lb1.grid(row=0,column=0, columnspan=2)
  
  lbeno = Label(searchForm,text="Employee Id",bg='black',fg='yellow',font=('verdana',10,'bold'))
  lbeno.grid(row=1,column=0,pady=10,sticky=W,padx=4)
  Button(searchForm,text="Find",command=lambda:findRecord(eno,en,da,sl,searchForm)).grid(row=1,column=2)
  enteno = Entry(searchForm,textvariable=eno,width=20,bg='yellow',fg='red',font=('verdana',10,'bold'))
  enteno.grid(row=1,column=1,pady=10,sticky=W)
  
  lbname = Label(searchForm,text="Employee Name",bg='black',fg='yellow',font=('verdana',10,'bold'))
  lbname.grid(row=2,column=0,pady=10,padx=4,sticky=W)
  
  entname = Entry(searchForm,textvariable=en,width=20,bg='yellow',fg='red',font=('verdana',10,'bold'))
  entname.grid(row=2,column=1,pady=10,sticky=W)
  
  lbdept = Label(searchForm,text="Employee Address",bg='black',fg='yellow',font=('verdana',10,'bold'))
  lbdept.grid(row=3,column=0,pady=10,sticky=W,padx=4)
  
  entdept = Entry(searchForm,textvariable=da,width=20,bg='yellow',fg='red',font=('verdana',10,'bold'))
  entdept.grid(row=3,column=1,pady=10,sticky=W)
  
  lbsalary = Label(searchForm,text="Employee Salary",bg='black',fg='yellow',font=('verdana',10,'bold'))
  lbsalary.grid(row=4,column=0,pady=10,sticky=W,padx=4)
  
  entsalary = Entry(searchForm,textvariable=sl,width=20,bg='yellow',fg='red',font=('verdana',10,'bold'))
  entsalary.grid(row=4,column=1,pady=10,sticky=W)

  Button(searchForm,text="Close Me!",command=searchForm.destroy,width=20,bg='black',fg='white',font=('verdana',10,'bold')).grid(row=6,column=0,columnspan=2,pady=5)
'''
def about():
        aboutmessage=''' 
        Author : Sujith Kumar Reddy
        E-Mail : vtu15960@veltehch.edu.in
        url    : sujithreddychevvu.blogspot.com'''
                
        messagebox.showinfo('...::::About US',aboutmessage)
        
def enterinfo():
  txtpayslip.delete("1.0",END)
  txtpayslip.insert(END,"\t      Pay Slip\n\n")
  txtpayslip.insert(END,"Name :\t\t"+Name.get()+"\n\n")
  txtpayslip.insert(END,"Address :\t\t"+Address.get()+"\n\n")
  txtpayslip.insert(END,"Employer :\t\t"+Employer.get()+"\n\n")
  txtpayslip.insert(END,"Employee Id :\t\t"+EmployeeId.get()+"\n\n")
  txtpayslip.insert(END,"Hours Worked :\t\t"+HoursWorked.get()+"\n\n")
  txtpayslip.insert(END,"Net Payable :\t\t"+NetPayable.get()+"\n\n")
  txtpayslip.insert(END,"Wages per hour :\t\t"+wageshour.get()+"\n\n")
  txtpayslip.insert(END,"Tax Paid :\t\t"+Taxable.get()+"\n\n")
  txtpayslip.insert(END,"Payable :\t\t"+Payable.get()+"\n\n") 
def weeklywages():
  txtpayslip.delete("1.0",END)
  hoursworkedperweek=float(HoursWorked.get())
  wagesperhours=float(wageshour.get())
  paydue=wagesperhours*hoursworkedperweek
  paymentdue="INR",str('%.2f'%(paydue))
  Payable.set(paymentdue)
  tax=paydue*0.2
  taxable="INR",str('%.2f'%(tax))
  Taxable.set(taxable)
  netpay=paydue-tax
  netpays="INR",str('%.2f'%(netpay))
  NetPayable.set(netpays)
  
  if hoursworkedperweek > 40:
    overtimehours=(hoursworkedperweek-40)+wagesperhours*1.5
    overtime="INR",str('%.2f'%(overtimehours))
    OverTimeBonus.set(overtime)
  elif hoursworkedperweek<=40:
    overtimepay=(hoursworkedperweek-40)+wagesperhours*1.5
    overtimehrs="INR",str('%.2f'%(overtimepay))
    OverTimeBonus.set(overtimehrs)  
  return  
    
#=============================== Variables ========================================================
Name=StringVar()
Address=StringVar()
HoursWorked=StringVar()
wageshour=StringVar()
Payable=StringVar()
Taxable=StringVar()
NetPayable=StringVar()
GrossPayable=StringVar()
OverTimeBonus=StringVar()
Employer=StringVar()
EmployeeId=StringVar()
TimeOfOrder=StringVar()
DateOfOrder=StringVar()
 
DateOfOrder.set(time.strftime("%d/%m/%Y"))
 
#================================ Label Widget =================================================
 
lblName=Label(fla,text="Name",font=('arial',16,'bold'),bd=20,fg="black",bg="light yellow").grid(row=0,column=0)
lblAddress=Label(fla,text="Address",font=('arial',16,'bold'),bd=20,fg="black",bg="light yellow").grid(row=0,column=2)
lblEmployer=Label(fla,text="Employer",font=('arial',16,'bold'),bd=20,fg="black",bg="light yellow").grid(row=1,column=0)
lblEmployeeId=Label(fla,text="Employee Id",font=('arial',16,'bold'),bd=20,fg="black",bg="light yellow").grid(row=1,column=2)
lblHoursWorked=Label(fla,text="Hours Worked",font=('arial',16,'bold'),bd=20,fg="black",bg="light yellow").grid(row=2,column=0)
lblHourlyRate=Label(fla,text="Hourly Rate",font=('arial',16,'bold'),bd=20,fg="black",bg="light yellow").grid(row=2,column=2)
lblTax=Label(fla,text="Tax",font=('arial',16,'bold'),bd=20,anchor='w',fg="black",bg="light yellow").grid(row=3,column=0)
lblOverTime=Label(fla,text="OverTime",font=('arial',16,'bold'),bd=20,fg="black",bg="light yellow").grid(row=3,column=2)
lblGrossPay=Label(fla,text="GrossPay",font=('arial',16,'bold'),bd=20,fg="black",bg="light yellow").grid(row=4,column=0)
lblNetPay=Label(fla,text="Net Pay",font=('arial',16,'bold'),bd=20,fg="black",bg="light yellow").grid(row=4,column=2)
 
#=============================== Entry Widget =================================================
 
etxname=Entry(fla,textvariable=Name,font=('arial',16,'bold'),bd=16,width=11,justify='left')
etxname.grid(row=0,column=1)
 
etxaddress=Entry(fla,textvariable=Address,font=('arial',16,'bold'),bd=16,width=11,justify='left')
etxaddress.grid(row=0,column=3)
 
etxemployer=Entry(fla,textvariable=Employer,font=('arial',16,'bold'),bd=16,width=11,justify='left')
etxemployer.grid(row=1,column=1)
 
etxhoursworked=Entry(fla,textvariable=HoursWorked,font=('arial',16,'bold'),bd=16,width=11,justify='left')
etxhoursworked.grid(row=2,column=1)
 
etxwagesperhours=Entry(fla,textvariable=wageshour,font=('arial',16,'bold'),bd=16,width=11,justify='left')
etxwagesperhours.grid(row=2,column=3)
 
etxnin=Entry(fla,textvariable=EmployeeId,font=('arial',16,'bold'),bd=16,width=11,justify='left')
etxnin.grid(row=1,column=3)
 
etxgrosspay=Entry(fla,textvariable=Payable,font=('arial',16,'bold'),bd=16,width=11,justify='left')
etxgrosspay.grid(row=4,column=1)
 
etxnetpay=Entry(fla,textvariable=NetPayable,font=('arial',16,'bold'),bd=16,width=11,justify='left')
etxnetpay.grid(row=4,column=3)
 
etxtax=Entry(fla,textvariable=Taxable,font=('arial',16,'bold'),bd=16,width=11,justify='left')
etxtax.grid(row=3,column=1)
 
etxovertime=Entry(fla,textvariable=OverTimeBonus,font=('arial',16,'bold'),bd=16,width=11,justify='left')
etxovertime.grid(row=3,column=3)
 
#=============================== Text Widget ============================================================
 
payslip=Label(f2,textvariable=DateOfOrder,font=('arial',17,'bold'),fg="black",bg="light yellow").grid(row=0,column=0)
txtpayslip=Text(f2,height=20,width=30,bd=14,font=('arial',10,'bold'),fg="black",bg="light yellow")
txtpayslip.grid(row=1,column=0)
 
#=============================== buttons ===============================================================

btnsalary=Button(flb,text='Weekly Salary',padx=15,pady=15,bd=6,font=('arial',10,'bold'),width=6,fg="black",bg="orange",command=weeklywages).grid(row=0,column=0)
 
#btnreset=Button(flb,text='Reset',padx=15,pady=15,bd=6,font=('arial',10,'bold'),width=6,command=reset,fg="black",bg="powder blue").grid(row=0,column=1)
 
#btnpayslip=Button(flb,text='View Payslip',padx=15,pady=15,bd=6,font=('arial',10,'bold'),width=6,command=enterinfo,fg="black",bg="powder blue").grid(row=0,column=2)
 
btnexit=Button(flb,text='Exit System',padx=15,pady=15,bd=6,font=('arial',10,'bold'),width=6,command=exit,fg="black",bg="orange").grid(row=0,column=3)

#btnsave=Button(flb,text='Save',padx=15,pady=15,bd=6,font=('arial',10,'bold'),width=6,command=saoveinfo,fg="black",bg="powder blue").grid(row=0,column=4)

btnexport=Button(flb,text='Export',padx=15,pady=15,bd=6,font=('arial',10,'bold'),width=6,command=export,fg="black",bg="orange").grid(row=0,column=5)

#btnabout=Button(flb,text='Find',padx=15,pady=15,bd=6,font=('arial',10,'bold'),width=6,command=find,fg="black",bg="orange").grid(row=0,column=6)

btnabout=Button(flb,text='About',padx=15,pady=15,bd=6,font=('arial',10,'bold'),width=6,command=about,fg="black",bg="orange").grid(row=0,column=7)

root.mainloop()
